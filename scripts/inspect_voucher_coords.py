import pandas as pd
import openpyxl

file_path = "c:\\Users\\User\\OneDrive\\Escritorio\\PROJETOS\\07_TET\\AppyTET\\data_source\\CONFIRMACIONES 2025.xlsx"

try:
    # Use openpyxl directly to get cell coordinates easily
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'Voucher' in wb.sheetnames:
        ws = wb['Voucher']
        print(f"--- Content of sheet 'Voucher' ---")
        for row in ws.iter_rows(min_row=1, max_row=40, max_col=10):
            for cell in row:
                if cell.value:
                    print(f"{cell.coordinate}: {cell.value}")
    else:
        print("'Voucher' sheet not found.")
        print(f"Available sheets: {wb.sheetnames}")

except Exception as e:
    print(f"Error: {e}")
